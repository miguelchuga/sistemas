# -*- coding: utf-8 -*-
from tempfile import NamedTemporaryFile
from openpyxl import Workbook
from openpyxl.styles import Color, Fill, Font, Alignment
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side

import base64
from openerp import models, fields

class GenerateStockWizard(models.Model):
    
    _name = "libro_ventas.generate_ventas"
    data = fields.Binary('File', readonly=True)
    name = fields.Char('File Name', readonly=True)
    state = fields.Selection([('choose', 'choose'),
                              ('get', 'get')], default='choose')
    period_id=fields.Many2one('account.period',string=' ')

    def generate_file(self, cr, uid, ids, context=None):

        if context is None:
            context = {}

        period_id_code = str(self.browse(cr, uid, ids)[0].period_id.code)

        user = self.pool.get('res.users').browse(cr, uid, uid, context=context)

        company_id = user.company_id.id

        #Totales del reporte 
        Tbienes_gravado_local = 0
        Tbienes_nogravado_local = 0
        Tservicios_gravado_local = 0                     
        Tservicios_nogravado_local = 0
        Tiva = 0
        Ttotal = 0

        this = self.browse(cr, uid, ids)[0]
        fileobj = NamedTemporaryFile('w+b')
        xlsfile = fileobj.name
        fileobj.close()
        thin_border = Border(left=Side(style='thin'), 
                      right=Side(style='thin'), 
                      top=Side(style='thin'), 
                      bottom=Side(style='thin'))

        wb = Workbook()

        ws = wb.active

        ws.title = "ventas"
        ws['A1'].value = "SUPERINTENDENCIA DE ADMINISTRACIÓN TRIBUTARIA"
        ws['A2'].value = "ASISTE LIBROS"
        ws['A3'].value = "LIBRO DE VENTAS Y SERVICIOS PRESTADOS"
        ws['A4'].value = user.company_id.name
        ws.merge_cells('A1:P1') 
        ws.merge_cells('A2:P2') 
        ws.merge_cells('A3:P3') 
        ws.merge_cells('A4:P4')
        for r in ws.iter_rows('A1:P4'):
            for c in r:
                c.font = c.font.copy(size=12, bold=True, italic=True)
                c.alignment = Alignment(horizontal='center')

#Primera fila de título de columna.
        ws['H6'].value = "Ventas"
        ws['I6'].value = "Ventas"
        ws['J6'].value = "Servicios"
        ws['K6'].value = "Servicios"
        ws['L6'].value = "Número de"
        ws['M6'].value = "Valor de la"

#Primera fila de título de columna.
        ws['B7'].value = "Serie del"
        ws['C7'].value = "Número del"
        ws['D7'].value = "Fecha del"
        ws['G7'].value = "Estado del"
        ws['H7'].value = "Gravadas"
        ws['I7'].value = "Exentas"
        ws['J7'].value = "Gravados"
        ws['K7'].value = "Exentos"
        ws['L7'].value = "Tipo de"
        ws['M7'].value = "Constancia de"
        ws['N7'].value = "Constancia de"

#Última fila de título de columna.
        ws['A8'].value = "Documento"
        ws['B8'].value = "Documento"        
        ws['C8'].value = "Documento"
        ws['D8'].value = "Documento" 
        ws['E8'].value = "NIT"
        ws['F8'].value = "Nombre"
        ws['G8'].value = "Documento"
        ws['H8'].value = "Operación Local"
        ws['I8'].value = "Operación Local"
        ws['J8'].value = "Operación Local"
        ws['K8'].value = "Operación Local"             
        ws['L8'].value = "Constancia"
        ws['M8'].value = "Retención del IVA"
        ws['N8'].value = "Retención del IVA"
        ws['O8'].value = "IVA"
        ws['P8'].value = "Total"

#Locales.
        #ws.merge_cells('C6:F6') #Datos del Documento.
        #ws.merge_cells('G6:J6') #Datos vendor o Prestador de servicios.
        #ws.merge_cells('K6:N6') #local.
        #ws.merge_cells('O6:R6') #local.
        #ws.merge_cells('K7:L7') #gravadas.
        #ws.merge_cells('M7:N7') #exentas.



        sql = """
           SELECT id, tipo_transaccion, documento, serie_documento, documento_partner, constancia,
                   fecha, nit, nombre, estado, diario, tipo_gasto, local, gravado,
                   asiste_libro, id_diario, serie_venta, tipo_venta, tipo_constancia, origin, code,
                   COALESCE(bienes_gravado_local,0) bienes_gravado_local, COALESCE(servicios_gravado_local,0) servicios_gravado_local, 0 bienes_nogravado_local, 0 servicios_nogravado_local,
                   COALESCE(iva,0) iva, COALESCE(total,0) total, COALESCE(valor_constancia,0) valor_constancia,company_id
            FROM "MC_ventas" Where company_id = %s and code = %s Order by fecha
            """

        cr.execute(sql,(company_id,period_id_code,))

        row = 9

        doc_count = 0

        #Lee cada registro.        
        for query_line in cr.dictfetchall():

            fs = query_line['fecha'].split('-')
            _fecha=((fs[2])+"-"+fs[1]+"-"+fs[0])

            #Documento.    
            ws.cell(row=row, column=1).value  = query_line['tipo_venta']
            #Serie del documento.
            ws.cell(row=row, column=2).value  = query_line['serie_venta'] 
            #Número del documento.
            ws.cell(row=row, column=3).value  = query_line['documento_partner']
            #Fecha del documento.
            ws.cell(row=row, column=4).value  =_fecha
            #NIT del cliente/proveedor.
            ws.cell(row=row, column=5).value  = query_line['nit']
            #Nombre del cliente/proveedor.
            ws.cell(row=row, column=6).value  = query_line['nombre']
            #Estado del documento.
            ws.cell(row=row, column=7).value  = query_line['estado']
            #Valor Ventas Gravadas operación local. 
            ws.cell(row=row, column=8).value =  query_line['bienes_gravado_local']
            #Valor Ventas Exentas operación local.
            ws.cell(row=row, column=9).value =  query_line['bienes_nogravado_local']
            #Valor Servicios Gravados operación local. 
            ws.cell(row=row, column=10).value =  query_line['servicios_gravado_local']
            #Valor Serviciso Exentos operación local.
            ws.cell(row=row, column=11).value =  query_line['servicios_nogravado_local']
            #Tipo de constancia.
            ws.cell(row=row, column=12).value  = query_line['tipo_constancia']            
            #Número de la constancia de retención del IVA.
            ws.cell(row=row, column=13).value  = query_line['constancia']
            #Valor de la constancia de retención del IVA. 
            ws.cell(row=row, column=14).value  = query_line['valor_constancia']
            #IVA.
            ws.cell(row=row, column=15).value = query_line['iva']
            #Total.
            ws.cell(row=row, column=16).value = query_line['total']          
            
            #Totaliza las columnas.
            Tbienes_gravado_local         += query_line['bienes_gravado_local'] 
            Tbienes_nogravado_local       += query_line['bienes_nogravado_local']            
            Tservicios_gravado_local      += query_line['servicios_gravado_local'] 
            Tservicios_nogravado_local    += query_line['servicios_nogravado_local']  
            Tiva   += query_line['iva']
            Ttotal += query_line['total']
            
            doc_count += 1
            row += 1

        # Totales del reporte.
        ws.cell(row=row, column=5).value  = 'Total: ' 
        ws.cell(row=row, column=6).value  = doc_count
        ws.cell(row=row, column=7).value  = 'Totales en Q.' 
        ws.cell(row=row, column=8).value  = Tbienes_gravado_local
        ws.cell(row=row, column=9).value  = Tbienes_nogravado_local
        ws.cell(row=row, column=10).value  = Tservicios_gravado_local
        ws.cell(row=row, column=11).value  = Tservicios_nogravado_local            
        ws.cell(row=row, column=15).value = Tiva
        ws.cell(row=row, column=16).value = Ttotal
        
        #Bold y centrado para los Totales.
        for c in range(06,06):
            ws.cell(row=row, column=c).border = thin_border
        for c in range(07,07):
            ws.cell(row=row, column=c).border = thin_border
        for c in range(8,17):
            ws.cell(row=row, column=c).border = thin_border
            
        #Aquí arriba hay que poner bold los totales (investigar).

        #Resumen del reporte.
        rrow = row + 4
                
        # Aquí falta el styl para que centre los titulos.
        rrow += 1
        cCell1 = ws.cell(row = rrow, column = 11)
        cCell2 = ws.cell(row = rrow, column = 17)
       
        for r in ws.iter_rows(cCell1.coordinate+':'+cCell2.coordinate):
            for c in r:
                c.font = c.font.copy(size=10, bold=True, italic=True)
        
        #Borde de todo el contenido o detalle.
        for r in range(9, row):
            for c in range(1,17):
                ws.cell(row=r, column=c).border = thin_border

        #Centrado y bold de los títulos.
        for r in ws.iter_rows('A6:P8'):
            for c in r:
                c.font = c.font.copy(size=10, bold=True, italic=True)
                c.alignment = Alignment(horizontal='center')
                
        #Pone todas las celdas con font 10.
        cCell1 = ws.cell(row = row, column = 01)
        cCell2 = ws.cell(row = row, column = 16)

        for r in ws.iter_rows(cCell1.coordinate+':'+cCell2.coordinate):
            for c in r:
                c.font = c.font.copy(size=10)

        #subrayado.
        #1er subrayado.
        for r in ws.iter_rows('A8:A8'):
            for c in r:
                c.border = thin_border
        #2do subrayado.
        for r in ws.iter_rows('B7:B8'):
            for c in r:
                c.border = thin_border
        #3er subrayado.
        for r in ws.iter_rows('C6:P8'):
            for c in r:
                c.border = thin_border
        
        #Amplía ancho de columna.
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 50
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 14
        ws.column_dimensions['J'].width = 16
        ws.column_dimensions['K'].width = 16
        ws.column_dimensions['L'].width = 16
        ws.column_dimensions['M'].width = 16
        ws.column_dimensions['N'].width = 16

        wb.save(filename=xlsfile)

        spreadsheet_file = open(xlsfile, "rb")
        binary_data = spreadsheet_file.read()
        spreadsheet_file.close()
        out = base64.b64encode(binary_data)

        self.write(cr, uid, ids, {
            'state': 'get',
            'name': "libro_ventas_spreadsheet.xlsx",
            'data': out
        }, context=context)

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'libro_ventas.generate_ventas',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': this.id,
            'views': [(False, 'form')],
            'target': 'new',
        }
